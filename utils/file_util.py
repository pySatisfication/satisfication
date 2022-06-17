import os
import openpyxl
from openpyxl.styles import Alignment

def num_to_title(n):
    res = ''
    while n > 0:
        res = (chr)((n-1)%26+65) + res
        n = int((n-1)/26)
    return res

def save_xlsx(file_name, content, tips=None, column_width=20):
    ex_file = openpyxl.Workbook()
    ex_file.create_sheet('指标校验')
    table = ex_file.active

    lineidx = 1
    column_num = 0
    for idx in range(len(content)):
        colidx = 1
        for item in content[idx]:
            table.cell(lineidx, colidx, item)
            colidx += 1
            column_num += 1
        lineidx += 1

    for item in range(1, column_num+1):
        if item == 1:
            width = column_width*2
        else:
            width = column_width
        table.column_dimensions[num_to_title(item)].width = width
    
    ex_file.save(file_name)

def get_sub_dirs(path, filters=''):
    res = []
    t_files = os.listdir(path)
    for item in t_files:
        if os.path.isdir(os.path.join(path, item)):
            res.append(item)
    return res

def get_all_files(path, filters=''):
    res = []
    temp = os.listdir(path)
    for t in temp:
        new_path = os.path.join(path, t)
        if os.path.isfile(new_path):
            if filters in t:
                res.append(new_path)
        else:
            res.extend(get_all_files(new_path, filters))
    return res

if __name__ == '__main__':
    #print(get_files('/Users/tyree/Downloads/LH2207/4/2021', '_6_5'))
    #print(num_to_title(27))
    print(get_sub_dirs('/Users/tyree/Downloads/LH'))
